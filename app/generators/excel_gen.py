"""
Excel generator — 6-tab workbook following SOP v4.
Tabs: Executive Summary, Traffic & Sales, Advertising, Profit & Loss,
      Inventory & BSR, Account Health.
"""
import io
from dataclasses import asdict
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers as xl_numbers
)
from openpyxl.utils import get_column_letter

from app.engine.kpis import (
    AccountSummary, AdPerformance, KPI_LABELS, KPI_TYPE,
    ACCOUNT_FIELDS, ADS_FIELDS,
)
from app.engine.deltas import build_deltas

# ── Capybaras brand colours ────────────────────────────────────────────────
C_ORANGE = "E84000"
C_DARK = "1A1A1A"
C_LIGHT = "FAFAF8"
C_GREEN = "16803C"
C_RED = "DC2626"
C_BORDER = "E8E5E0"
C_HEADER_BG = "F5F3F0"
C_WHITE = "FFFFFF"

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color=C_DARK, size=10) -> Font:
    return Font(name="Calibri", bold=bold, color=color, size=size)

def _border() -> Border:
    thin = Side(style="thin", color=C_BORDER)
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _fmt_value(val, kpi_type: str, currency: str = "USD") -> str:
    if val is None:
        return "N/A"
    sym = "$" if currency == "USD" else ("MX$" if currency == "MXN" else "$")
    if kpi_type == "currency":
        return f"{sym}{val:,.2f}"
    if kpi_type == "percent":
        return f"{val:.2f}%"
    if kpi_type == "ratio":
        return f"{val:.2f}x"
    # number
    if isinstance(val, int) or val == int(val):
        return f"{int(val):,}"
    return f"{val:,.2f}"


def _write_kpi_table(
    ws,
    start_row: int,
    fields: list[str],
    current_obj,
    prior_obj,
    deltas: dict,
    currency: str = "USD",
) -> int:
    """Write a 5-column KPI comparison table. Returns last used row."""

    # Header row
    headers = ["KPI", f"Current", "Prior Year", "YoY Change", "Direction"]
    widths = [22, 16, 16, 14, 12]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=start_row, column=ci, value=h)
        cell.font = _font(bold=True, color=C_WHITE)
        cell.fill = _fill(C_DARK)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border()
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[start_row].height = 18
    r = start_row + 1

    for field in fields:
        label = KPI_LABELS.get(field, field)
        ktype = KPI_TYPE.get(field, "number")
        curr_val = getattr(current_obj, field) if current_obj else None
        prior_val = getattr(prior_obj, field) if prior_obj else None
        delta = deltas.get(field, {})

        curr_str = _fmt_value(curr_val, ktype, currency)
        prior_str = _fmt_value(prior_val, ktype, currency)
        delta_str = delta.get("display", "N/A")
        is_good = delta.get("is_good")
        direction = delta.get("direction", "neutral")

        row_bg = C_WHITE if r % 2 == 0 else C_LIGHT
        arrow = "▲" if direction == "up" else ("▼" if direction == "down" else "—")
        if is_good is True:
            delta_color = C_GREEN
        elif is_good is False:
            delta_color = C_RED
        else:
            delta_color = C_DARK

        data = [label, curr_str, prior_str, delta_str, arrow]
        for ci, val in enumerate(data, 1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.fill = _fill(row_bg)
            cell.border = _border()
            cell.alignment = Alignment(horizontal="center" if ci > 1 else "left",
                                       vertical="center")
            if ci == 4:
                cell.font = _font(bold=True, color=delta_color)
            elif ci == 5:
                cell.font = _font(bold=True,
                                  color=delta_color if is_good is not None else "888888")
            else:
                cell.font = _font()

        ws.row_dimensions[r].height = 16
        r += 1

    return r


def _section_header(ws, row: int, text: str, col_span: int = 5) -> None:
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=col_span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = _font(bold=True, color=C_WHITE, size=11)
    cell.fill = _fill(C_ORANGE)
    cell.alignment = Alignment(horizontal="left", vertical="center",
                                indent=1)
    ws.row_dimensions[row].height = 20


def generate(
    account_name: str,
    period_current: str,
    period_prior: str,
    current_account: AccountSummary,
    prior_account: AccountSummary,
    current_ads: AdPerformance,
    prior_ads: AdPerformance,
    week_number: Optional[int] = None,
    currency: str = "USD",
) -> bytes:
    """Return the 6-tab Excel workbook as bytes."""

    deltas = build_deltas(current_account, prior_account, current_ads, prior_ads)
    wb = Workbook()
    wb.remove(wb.active)

    # ── 1. Executive Summary ───────────────────────────────────────────────
    ws = wb.create_sheet("Executive Summary")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    ws["A1"] = account_name
    ws["A1"].font = _font(bold=True, color=C_WHITE, size=14)
    ws["A1"].fill = _fill(C_ORANGE)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:E2")
    ws["A2"] = f"Weekly YoY Report  ·  {period_current}  vs  {period_prior}"
    ws["A2"].font = _font(color="888888")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    r = 4
    _section_header(ws, r, "ACCOUNT SUMMARY — TRAFFIC & SALES")
    r += 1
    r = _write_kpi_table(ws, r, ACCOUNT_FIELDS, current_account, prior_account,
                         deltas, currency)

    r += 1
    _section_header(ws, r, "ADVERTISING PERFORMANCE SUMMARY")
    r += 1
    _write_kpi_table(ws, r, ADS_FIELDS, current_ads, prior_ads, deltas, currency)

    # ── 2. Traffic & Sales ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Traffic & Sales")
    ws2.sheet_view.showGridLines = False
    ws2.merge_cells("A1:E1")
    ws2["A1"] = "Account Summary — Traffic & Sales"
    ws2["A1"].font = _font(bold=True, color=C_WHITE, size=12)
    ws2["A1"].fill = _fill(C_ORANGE)
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 25
    _write_kpi_table(ws2, 3, ACCOUNT_FIELDS, current_account, prior_account,
                     deltas, currency)

    # ── 3. Advertising ────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Advertising")
    ws3.sheet_view.showGridLines = False
    ws3.merge_cells("A1:E1")
    ws3["A1"] = "Advertising Performance Summary"
    ws3["A1"].font = _font(bold=True, color=C_WHITE, size=12)
    ws3["A1"].fill = _fill(C_ORANGE)
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 25
    _write_kpi_table(ws3, 3, ADS_FIELDS, current_ads, prior_ads, deltas, currency)

    # ── 4. Profit & Loss ─────────────────────────────────────────────────
    ws4 = wb.create_sheet("Profit & Loss")
    ws4.sheet_view.showGridLines = False
    ws4.merge_cells("A1:C1")
    ws4["A1"] = "Profit & Loss"
    ws4["A1"].font = _font(bold=True, color=C_WHITE, size=12)
    ws4["A1"].fill = _fill(C_ORANGE)
    ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 25

    sym = "$" if currency == "USD" else "MX$"
    pl_rows = [
        ("", f"{period_current}", f"{period_prior}"),
        ("Gross Revenue (B2C)", _fmt_value(current_account.revenue, "currency", currency),
         _fmt_value(prior_account.revenue, "currency", currency)),
        ("Ad Spend", _fmt_value(current_ads.ad_spend if current_ads else None, "currency", currency),
         _fmt_value(prior_ads.ad_spend if prior_ads else None, "currency", currency)),
        ("Ad Spend % of Revenue",
         _fmt_value(current_ads.tacos_pct if current_ads else None, "percent"),
         _fmt_value(prior_ads.tacos_pct if prior_ads else None, "percent")),
        ("", "", ""),
        ("— COGs (add manually)", "", ""),
        ("— FBA Fees (add manually)", "", ""),
        ("— Other Expenses (add manually)", "", ""),
        ("", "", ""),
        ("Net Profit (manual)", "", ""),
    ]
    for ri, row_data in enumerate(pl_rows, start=3):
        for ci, val in enumerate(row_data, 1):
            cell = ws4.cell(row=ri, column=ci, value=val)
            cell.border = _border()
            cell.alignment = Alignment(horizontal="center" if ci > 1 else "left",
                                       vertical="center", indent=1 if ci == 1 else 0)
            is_header = ri == 3
            cell.font = _font(bold=is_header)
            cell.fill = _fill(C_DARK if is_header else (C_WHITE if ri % 2 == 0 else C_LIGHT))
            if is_header:
                cell.font = _font(bold=True, color=C_WHITE)
        ws4.row_dimensions[ri].height = 16

    for ci, w in enumerate([28, 18, 18], 1):
        ws4.column_dimensions[get_column_letter(ci)].width = w

    # Note cell
    note_row = 3 + len(pl_rows) + 1
    ws4.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=3)
    ws4.cell(row=note_row, column=1,
             value="Nota: No incluye Estimated Payout. Completar COGs y fees manualmente.").font = Font(
        italic=True, color="888888", size=9)

    # ── 5. Inventory & BSR ───────────────────────────────────────────────
    ws5 = wb.create_sheet("Inventory & BSR")
    ws5.sheet_view.showGridLines = False
    ws5.merge_cells("A1:C1")
    ws5["A1"] = "Inventory & BSR"
    ws5["A1"].font = _font(bold=True, color=C_WHITE, size=12)
    ws5["A1"].fill = _fill(C_ORANGE)
    ws5["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws5.row_dimensions[1].height = 25
    ws5.merge_cells("A3:C3")
    ws5["A3"] = ("Datos de Inventory & BSR no están disponibles en los archivos fuente. "
                 "Amazon no archiva BSR histórico ni inventory snapshot. "
                 "Agregar manualmente si disponible.")
    ws5["A3"].font = Font(italic=True, color="888888", size=9)
    ws5["A3"].alignment = Alignment(wrap_text=True)
    ws5.row_dimensions[3].height = 50

    # ── 6. Account Health ────────────────────────────────────────────────
    ws6 = wb.create_sheet("Account Health")
    ws6.sheet_view.showGridLines = False
    ws6.merge_cells("A1:C1")
    ws6["A1"] = "Account Health"
    ws6["A1"].font = _font(bold=True, color=C_WHITE, size=12)
    ws6["A1"].fill = _fill(C_ORANGE)
    ws6["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws6.row_dimensions[1].height = 25

    health_metrics = [
        "Order Defect Rate", "Late Shipment Rate", "Cancellation Rate",
        "A-to-Z Claims", "Policy Violations", "Customer Service Dissatisfaction",
    ]
    ws6.cell(row=3, column=1, value="Metric").font = _font(bold=True)
    ws6.cell(row=3, column=2, value=period_current).font = _font(bold=True)
    ws6.cell(row=3, column=3, value=f"{period_prior} (N/A)").font = _font(bold=True, color="888888")
    for ci in range(1, 4):
        ws6.cell(row=3, column=ci).fill = _fill(C_DARK)
        ws6.cell(row=3, column=ci).font = _font(bold=True, color=C_WHITE)
        ws6.cell(row=3, column=ci).border = _border()
    for ri, m in enumerate(health_metrics, start=4):
        ws6.cell(row=ri, column=1, value=m).border = _border()
        ws6.cell(row=ri, column=2, value="— (agregar manualmente)").border = _border()
        ws6.cell(row=ri, column=3, value="N/A").border = _border()
        ws6.cell(row=ri, column=3).font = _font(color="888888")
        fill = _fill(C_WHITE if ri % 2 == 0 else C_LIGHT)
        for ci in range(1, 4):
            ws6.cell(row=ri, column=ci).fill = fill
    for ci, w in enumerate([30, 20, 20], 1):
        ws6.column_dimensions[get_column_letter(ci)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_filename(account_name: str, week_number: Optional[int],
                   period_current: str, period_prior: str) -> str:
    wk = f"Week {week_number:02d}" if week_number else "Weekly"
    clean = account_name.replace("/", "-")
    return f"{clean} - {wk} - {period_current} vs {period_prior}.xlsx"
