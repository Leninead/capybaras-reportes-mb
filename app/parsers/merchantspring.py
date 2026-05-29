"""
Parser for the current-period file (File A).

Supports three formats in priority order:
  1. Capybaras weekly_report Excel — has '▶ CUENTA TOTAL' aggregate row + Advertising section.
  2. Generic MerchantSpring / summary Excel — single header row, aggregate values.
  3. Amazon Business Report CSV — falls through to business_report.parse().
"""
import csv
import io
import logging
from typing import Optional
from .utils import parse_number, find_col
from app.engine.kpis import AccountSummary, AdPerformance, ParsedData

log = logging.getLogger(__name__)

# ──────────────────────────────────────────────────────────────────────────────
# Format 1: Capybaras weekly_report Excel
# ──────────────────────────────────────────────────────────────────────────────

def _parse_capybaras_weekly(ws_rows: list[list]) -> Optional[ParsedData]:
    """
    Parse the Capybaras weekly_report Excel format.
    Header structure (3 rows):
      Row N  : group labels (SALES, UNITS, SESSIONS, CVR, BUY BOX, AD SALES, AD SPEND, ACoS, TACoS)
      Row N+1: sub-labels  (Esta semana, Semana anterior, Variación %, ...)
      Row N+2: CUENTA TOTAL data row

    Advertising section comes later in the sheet:
      Row: Impressions, Clicks, CTR, CPC, Spend, Sales, ACoS, Orders
      Row: values
    """
    flat = [[str(c).strip() if c is not None else "" for c in row] for row in ws_rows]

    # Find the CUENTA TOTAL row
    cuenta_idx = None
    for i, row in enumerate(flat):
        if any("CUENTA TOTAL" in cell or "CUENTA\xa0TOTAL" in cell for cell in row):
            cuenta_idx = i
            break

    if cuenta_idx is None:
        return None

    log.info("MerchantSpring: detected Capybaras weekly_report format (CUENTA TOTAL at row %d)", cuenta_idx)

    # Build column-name map from the two header rows above CUENTA TOTAL
    # Row cuenta_idx-2: group labels
    # Row cuenta_idx-1: sub-labels
    group_row = flat[cuenta_idx - 2] if cuenta_idx >= 2 else []
    sub_row = flat[cuenta_idx - 1] if cuenta_idx >= 1 else []
    data_row = flat[cuenta_idx]

    # Walk columns, carrying the last non-empty group label
    col_map: dict[str, int] = {}  # logical_name → column index
    last_group = ""
    for ci in range(len(data_row)):
        group = group_row[ci] if ci < len(group_row) else ""
        sub = sub_row[ci] if ci < len(sub_row) else ""

        if group:
            last_group = group.upper().strip()

        # Current-week data is the first sub-column after each group
        is_current = "ESTA" in sub.upper() or sub == ""

        mapping = {
            "SALES": ("revenue", is_current),
            "VENTAS": ("revenue", is_current),
            "UNITS": ("ordered_units", is_current),
            "UNIDADES": ("ordered_units", is_current),
            "SESSIONS": ("sessions", is_current),
            "SESIONES": ("sessions", is_current),
            "CVR": ("conv_pct", is_current),
            "BUY BOX": ("buybox_win_pct", True),  # single column
            "AD SALES": ("ad_sales", is_current),
            "AD SPEND": ("ad_spend", is_current),
            "ACOS": ("acos_pct", True),
            "TACOS": ("tacos_pct", True),
        }

        if last_group in mapping:
            logical, take = mapping[last_group]
            if take and logical not in col_map:
                col_map[logical] = ci

    def get(key: str) -> Optional[float]:
        idx = col_map.get(key)
        if idx is None:
            return None
        val = data_row[idx] if idx < len(data_row) else ""
        return parse_number(val)

    revenue = get("revenue")
    units = get("ordered_units")
    sessions = get("sessions")
    conv_pct = get("conv_pct")
    buybox = get("buybox_win_pct")
    ad_sales = get("ad_sales")
    ad_spend = get("ad_spend")
    acos_pct = get("acos_pct")
    tacos_pct = get("tacos_pct")

    if revenue is None and sessions is None:
        log.warning("MerchantSpring: CUENTA TOTAL row found but values could not be extracted.")
        return None

    # Derived account metrics
    avg_retail = (revenue / units) if (revenue and units and units > 0) else None
    conv_pct_calc = (units / sessions * 100) if (units and sessions and sessions > 0) else conv_pct

    # Find Advertising Overview section
    impressions = clicks = orders = cpc_val = None
    for i, row in enumerate(flat):
        if any("IMPRESSIONS" in c.upper() for c in row):
            # Next non-empty row is the values
            for val_row in flat[i + 1:]:
                if any(c for c in val_row):
                    headers_adv = row
                    hi_map = {h.upper().strip(): ci for ci, h in enumerate(headers_adv) if h}
                    def get_adv(name: str) -> Optional[float]:
                        idx = hi_map.get(name)
                        if idx is None:
                            return None
                        return parse_number(val_row[idx] if idx < len(val_row) else None)
                    impressions = get_adv("IMPRESSIONS")
                    clicks = get_adv("CLICKS")
                    orders = get_adv("ORDERS")
                    cpc_val = get_adv("CPC")
                    if ad_sales is None:
                        ad_sales = get_adv("SALES")
                    if ad_spend is None:
                        ad_spend = get_adv("SPEND")
                    break
            break

    # Recalculate derived ads metrics
    acos_pct = acos_pct or ((ad_spend / ad_sales * 100) if (ad_spend and ad_sales and ad_sales > 0) else None)
    roas = (ad_sales / ad_spend) if (ad_sales and ad_spend and ad_spend > 0) else None
    cpc_calc = cpc_val or ((ad_spend / clicks) if (ad_spend and clicks and clicks > 0) else None)
    cpa = (ad_spend / orders) if (ad_spend and orders and orders > 0) else None
    conv_ads = (orders / clicks * 100) if (orders and clicks and clicks > 0) else None

    warnings = []
    if impressions is None:
        warnings.append("Sección Advertising Overview no encontrada en el reporte; métricas de impresiones/clicks no disponibles.")

    return ParsedData(
        account=AccountSummary(
            revenue=revenue,
            ordered_units=int(units) if units else None,
            page_views=None,  # WoW report doesn't show page views
            conv_pct=conv_pct_calc,
            sessions=int(sessions) if sessions else None,
            s_conv_pct=None,
            avg_retail=avg_retail,
            buybox_win_pct=buybox,
            mobile_sessions_pct=None,
        ),
        ads=AdPerformance(
            ad_sales=ad_sales,
            ad_spend=ad_spend,
            acos_pct=acos_pct,
            roas=roas,
            tacos_pct=tacos_pct,
            troas=None,
            impressions=int(impressions) if impressions else None,
            clicks=int(clicks) if clicks else None,
            orders=int(orders) if orders else None,
            units_ads=None,
            cpc=cpc_calc,
            cpa=cpa,
            conv_pct_ads=conv_ads,
            ntb_units=None,
        ),
        column_mapping={"merchantspring_capybaras": col_map},
        warnings=warnings,
    )


# ──────────────────────────────────────────────────────────────────────────────
# Format 2: Generic MerchantSpring / summary Excel
# ──────────────────────────────────────────────────────────────────────────────

GENERIC_ACCOUNT_ALIASES = {
    "revenue": [
        "Ordered Product Sales", "Revenue", "Sales", "Total Sales",
        "Gross Revenue", "Net Revenue",
    ],
    "ordered_units": ["Units Ordered", "Units", "Orders", "Total Units"],
    "page_views": ["Page Views", "Page Views - Total"],
    "sessions": ["Sessions", "Sessions - Total"],
    "conv_pct": [
        "Unit Session Percentage", "Conversion Rate", "CVR",
        "Conv. Rate", "Conv Rate",
    ],
    "s_conv_pct": ["Order Session Percentage", "S. Conv.", "S Conv"],
    "buybox_win_pct": [
        "Featured Offer (Buy Box) Percentage", "Buy Box %",
        "Buy Box Percentage", "Buybox Win",
    ],
    "mobile_sessions_pct": ["Mobile Sessions %", "Mobile Traffic %", "Mobile S."],
}

GENERIC_ADS_ALIASES = {
    "ad_sales": ["Ad Sales", "Advertising Sales", "Total Ad Sales"],
    "ad_spend": ["Ad Spend", "Spend", "Cost", "Advertising Spend"],
    "acos_pct": ["ACoS", "ACOS"],
    "roas": ["ROAS", "RoAS"],
    "tacos_pct": ["TACoS", "TACOS"],
    "impressions": ["Impressions"],
    "clicks": ["Clicks"],
    "orders": ["Orders", "Ad Orders"],
    "units_ads": ["Units", "Ad Units"],
    "cpc": ["CPC", "Cost Per Click"],
    "ntb_units": ["NTB Units", "New-to-brand Units"],
}


def _parse_generic_excel(ws_rows: list[list]) -> Optional[ParsedData]:
    """
    Try to parse a generic Excel summary with headers in row 0.
    Looks for aggregate row (Total / Account / Summary keyword).
    """
    flat = [[str(c).strip() if c is not None else "" for c in row] for row in ws_rows]
    if not flat:
        return None

    headers = flat[0]

    # Look for a data row: total/account row, or just the first non-empty data row
    data_row = None
    for row in flat[1:]:
        if any(c for c in row):
            # Prefer a "total" labeled row
            first = row[0].upper() if row[0] else ""
            if "TOTAL" in first or "ACCOUNT" in first or "SUMMARY" in first:
                data_row = row
                break
        if data_row is None and any(c for c in row):
            data_row = row

    if data_row is None:
        return None

    col_map: dict[str, int] = {}
    for key, aliases in {**GENERIC_ACCOUNT_ALIASES, **GENERIC_ADS_ALIASES}.items():
        for alias in aliases:
            for ci, h in enumerate(headers):
                if h.lower() == alias.lower():
                    col_map[key] = ci
                    break
            if key in col_map:
                break

    def get(key: str) -> Optional[float]:
        idx = col_map.get(key)
        if idx is None:
            return None
        val = data_row[idx] if idx < len(data_row) else ""
        return parse_number(val)

    revenue = get("revenue")
    if revenue is None:
        return None  # can't determine it's a valid summary sheet

    units = get("ordered_units")
    sessions = get("sessions")
    page_views = get("page_views")
    conv_pct = get("conv_pct")
    s_conv_pct = get("s_conv_pct")
    buybox = get("buybox_win_pct")
    mobile = get("mobile_sessions_pct")

    avg_retail = (revenue / units) if (revenue and units and units > 0) else None
    if conv_pct is None and units and sessions and sessions > 0:
        conv_pct = units / sessions * 100

    ad_sales = get("ad_sales")
    ad_spend = get("ad_spend")
    acos_pct = get("acos_pct")
    roas = get("roas")
    tacos_pct = get("tacos_pct")
    impressions = get("impressions")
    clicks = get("clicks")
    orders = get("orders")
    units_ads = get("units_ads")
    cpc = get("cpc")
    ntb_units = get("ntb_units")

    if acos_pct is None and ad_spend and ad_sales and ad_sales > 0:
        acos_pct = ad_spend / ad_sales * 100
    if roas is None and ad_sales and ad_spend and ad_spend > 0:
        roas = ad_sales / ad_spend
    cpa = (ad_spend / orders) if (ad_spend and orders and orders > 0) else None
    cpc = cpc or ((ad_spend / clicks) if (ad_spend and clicks and clicks > 0) else None)
    conv_ads = (orders / clicks * 100) if (orders and clicks and clicks > 0) else None

    return ParsedData(
        account=AccountSummary(
            revenue=revenue,
            ordered_units=int(units) if units else None,
            page_views=int(page_views) if page_views else None,
            conv_pct=conv_pct,
            sessions=int(sessions) if sessions else None,
            s_conv_pct=s_conv_pct,
            avg_retail=avg_retail,
            buybox_win_pct=buybox,
            mobile_sessions_pct=mobile,
        ),
        ads=AdPerformance(
            ad_sales=ad_sales,
            ad_spend=ad_spend,
            acos_pct=acos_pct,
            roas=roas,
            tacos_pct=tacos_pct,
            troas=None,
            impressions=int(impressions) if impressions else None,
            clicks=int(clicks) if clicks else None,
            orders=int(orders) if orders else None,
            units_ads=int(units_ads) if units_ads else None,
            cpc=cpc,
            cpa=cpa,
            conv_pct_ads=conv_ads,
            ntb_units=int(ntb_units) if ntb_units else None,
        ),
        column_mapping={"merchantspring_generic": col_map},
        warnings=["Formato MerchantSpring genérico detectado. Verificá todos los valores."],
    )


# ──────────────────────────────────────────────────────────────────────────────
# Public entry point
# ──────────────────────────────────────────────────────────────────────────────

def parse(content: bytes, filename: str = "") -> ParsedData:
    """
    Parse File A (current period). Auto-detects format:
      - Capybaras weekly_report Excel  (▶ CUENTA TOTAL)
      - Generic MerchantSpring Excel   (summary row)
      - Amazon Business Report CSV     (fallback)
    Raises ValueError with Spanish user-facing message on unrecoverable failure.
    """
    ext = (filename or "").lower()
    is_excel = ext.endswith(".xlsx") or ext.endswith(".xls")

    if is_excel:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)

        # Try each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))

            result = _parse_capybaras_weekly(rows)
            if result:
                log.info("Parsed File A as Capybaras weekly_report (sheet '%s')", sheet_name)
                return result

            result = _parse_generic_excel(rows)
            if result:
                log.info("Parsed File A as generic Excel summary (sheet '%s')", sheet_name)
                return result

        raise ValueError(
            "No pude reconocer el formato del archivo de período actual. "
            "Asegurate de subir el export de MerchantSpring (Excel) o el weekly_report generado por Capybaras."
        )

    else:
        # Try as Amazon Business Report CSV (same source data, different year)
        from . import business_report
        try:
            data = business_report.parse(content)
            data.warnings.insert(0, "Archivo A parseado como Amazon Business Report CSV.")
            return data
        except ValueError as exc:
            raise ValueError(
                f"No pude parsear el archivo de período actual. "
                f"Subí un Excel de MerchantSpring o el weekly_report de Capybaras. Detalle: {exc}"
            ) from exc
