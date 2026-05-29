"""
Parser for Amazon Ads Console Campaign Reports (SP + SB + SD combined).
Handles CSV and .xlsx. Aggregates all campaigns, then recalculates derived metrics.
"""
import csv
import io
import logging
from .utils import parse_number, find_col
from app.engine.kpis import AdPerformance, ParsedData

log = logging.getLogger(__name__)

# Amazon's column names vary by campaign type (SP vs SB vs SD) and attribution window.
ADS_COLUMN_ALIASES: dict[str, list[str]] = {
    "impressions": ["Impressions"],
    "clicks": ["Clicks"],
    "spend": [
        "Spend", "Cost", "Total Spend", "Ad Spend",
    ],
    "sales": [
        "7 Day Total Sales (#)", "7 Day Total Sales",
        "14 Day Total Sales (#)", "14 Day Total Sales",
        "Total Sales", "Sales", "Revenue", "Ad Sales",
    ],
    "orders": [
        "7 Day Total Orders (#)", "7 Day Total Orders",
        "14 Day Total Orders (#)", "14 Day Total Orders",
        "Total Orders", "Orders", "Ad Orders",
    ],
    "units": [
        "7 Day Total Units (#)", "7 Day Total Units",
        "14 Day Total Units (#)", "14 Day Total Units",
        "Total Units", "Units",
    ],
    "ntb_units": [
        "7 Day New-to-brand Units (#)", "7 Day New-to-brand Units",
        "14 Day New-to-brand Units (#)", "14 Day New-to-brand Units",
        "New-to-brand Units Ordered (#)", "New-to-brand Units Ordered",
        "NTB Units",
    ],
    "ntb_orders": [
        "7 Day New-to-brand Orders (#)", "7 Day New-to-brand Orders",
        "14 Day New-to-brand Orders (#)", "14 Day New-to-brand Orders",
        "New-to-brand Orders (#)", "New-to-brand Orders",
        "NTB Orders",
    ],
}

REQUIRED = ["impressions", "clicks", "spend"]


def _read_rows(content: bytes, filename: str) -> tuple[list[str], list[dict]]:
    """Return (headers, rows) from CSV or XLSX bytes."""
    ext = (filename or "").lower()
    if ext.endswith(".xlsx") or ext.endswith(".xls"):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            return [], []
        headers = [str(c) if c is not None else "" for c in all_rows[0]]
        rows = [
            {headers[i]: (str(row[i]) if row[i] is not None else "")
             for i in range(len(headers))}
            for row in all_rows[1:]
        ]
        return headers, rows
    else:
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        headers = list(reader.fieldnames or [])
        rows = list(reader)
        return headers, rows


def parse(content: bytes, filename: str = "") -> ParsedData:
    """
    Parse an Amazon Ads Campaign Report.
    Returns ParsedData with ads populated; account is None.
    """
    headers, rows = _read_rows(content, filename)

    col_map: dict[str, str] = {}
    for key, aliases in ADS_COLUMN_ALIASES.items():
        found = find_col(headers, aliases)
        if found:
            col_map[key] = found
            log.info("Ads Report: '%s' → column '%s'", key, found)

    missing_required = [k for k in REQUIRED if k not in col_map]
    if missing_required:
        cols_missing = ", ".join(f"'{ADS_COLUMN_ALIASES[k][0]}'" for k in missing_required)
        sample = ", ".join(f"'{h}'" for h in headers[:12])
        raise ValueError(
            f"No encontré las columnas requeridas en el Ads Report: {cols_missing}. "
            f"Columnas detectadas: {sample}{'...' if len(headers) > 12 else ''}. "
            "Asegurate de usar el reporte de Campaigns de Amazon Ads Console."
        )

    if not rows:
        raise ValueError("El Ads Report está vacío (sin filas de datos).")

    warnings: list[str] = []

    def sum_col(key: str) -> float:
        if key not in col_map:
            return 0.0
        return sum(parse_number(r.get(col_map[key])) or 0.0 for r in rows)

    impressions = sum_col("impressions")
    clicks = sum_col("clicks")
    spend = sum_col("spend")
    sales = sum_col("sales")
    orders = sum_col("orders")
    units = sum_col("units")
    ntb_units = sum_col("ntb_units") or sum_col("ntb_orders")

    if "sales" not in col_map:
        warnings.append(
            "Columna de ventas (Sales) no encontrada en Ads Report. "
            "ACOS, ROAS, TACOS y TROAS no podrán calcularse."
        )
    if "orders" not in col_map:
        warnings.append("Columna 'Orders' no encontrada en Ads Report. CPA y CONV no disponibles.")
    if "ntb_units" not in col_map and "ntb_orders" not in col_map:
        warnings.append("Columna NTB Units no encontrada en Ads Report.")

    # Recalculate all derived metrics from aggregated raw values
    acos_pct = (spend / sales * 100) if sales > 0 else None
    roas = (sales / spend) if spend > 0 else None
    cpc = (spend / clicks) if clicks > 0 else None
    cpa = (spend / orders) if orders > 0 else None
    conv_pct_ads = (orders / clicks * 100) if clicks > 0 else None

    return ParsedData(
        account=None,
        ads=AdPerformance(
            ad_sales=sales if sales > 0 else None,
            ad_spend=spend if spend > 0 else None,
            acos_pct=acos_pct,
            roas=roas,
            tacos_pct=None,   # requires total revenue — filled by caller
            troas=None,
            impressions=int(impressions),
            clicks=int(clicks),
            orders=int(orders) if orders > 0 else None,
            units_ads=int(units) if units > 0 else None,
            cpc=cpc,
            cpa=cpa,
            conv_pct_ads=conv_pct_ads,
            ntb_units=int(ntb_units) if ntb_units > 0 else None,
        ),
        column_mapping={"ads_report": col_map},
        warnings=warnings,
    )
